#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Jun 29 22:44:17 2024

@author: user
"""

import os
from langchain.tools import BaseTool
from rdkit import Chem, DataStructs
from rdkit.Chem import AllChem, rdMolDescriptors
from chemcrow.utils import *
from rdkit.Chem.SaltRemover import SaltRemover
from pydantic import ValidationError
import pandas as pd
from rdkit.Chem import Descriptors, QED
from rdkit.Chem.rdMolDescriptors import CalcTPSA
from rdkit.Chem.Crippen import MolLogP
import matplotlib.pyplot as plt
from rdkit.Chem import Draw
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from io import BytesIO
import re


class CSV2Structre_image(BaseTool):
    name = "CSV2Structre_image"
    description = "Input is the csv file, draw the molecular image based smiles, then returns the xlsx file including the image."

    def __init__(
        self,
    ):
        super(CSV2Structre_image, self).__init__()

    def _run(self, path_file: str) -> str:
        
        try:
            
            df = pd.read_csv(path_file)

            wb = Workbook()
            ws = wb.active
            nn =  len(df.columns) + 1
            pattern = re.compile(r'.*smiles.*', re.IGNORECASE)
            smiles_columns = [col for col in df.columns if pattern.match(col)]

            smiles_column = smiles_columns[0]
            if len(smiles_columns) != 1:
                raise ValueError("no column name includes 'smiles' ")

            for row in range(1, len(df) + 2):
                ws.row_dimensions[row].height = 250

            center_alignment = Alignment(horizontal='center', vertical='center')

            for c_idx, col_name in enumerate(df.columns):
                cell = ws.cell(row=1, column=c_idx + 1, value=col_name)
                # cell.alignment = center_alignment
                

                col_letter = chr(65 + c_idx)  
                if df.columns[c_idx] == smiles_column:
                    ws.column_dimensions[col_letter].width = 50
                else:
                    ws.column_dimensions[col_letter].width = 20

            for r_idx, row in df.iterrows():
                for c_idx, value in enumerate(row):
                    ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)
                    # cell.alignment = center_alignment

            img_column_index = len(df.columns) + 1
            ws.cell(row=1, column=img_column_index, value='structures')

            for idx, smiles in enumerate(df[smiles_column], start=2):
                molecule = Chem.MolFromSmiles(smiles)
                img = Draw.MolToImage(molecule, size=(400, 400))
                buffered = BytesIO()
                img.save(buffered, format="PNG")
                buffered.seek(0)
                
                image = Image(buffered)
                image.width = 300  
                image.height = 300 
                image.anchor = ws.cell(row=idx, column=nn).coordinate  # 将图片插入到B列
                ws.add_image(image)

            ws.column_dimensions[chr(65 + len(df.columns))].width = 50

            ws.row_dimensions[1].height = 20

            for row in ws.iter_rows(min_row=1, max_row=len(df) + 1, max_col=len(df.columns) + 2):
                for cell in row:
                    cell.alignment = center_alignment
                    
            filename = path_file[:-3] + 'xlsx'
            wb.save(filename)
            return filename
            
        except:
            return "The task execution is completed, please check the error in this step."
        
       
    async def _arun(self, path_file: str) -> str:
        """Use the tool asynchronously."""
        raise NotImplementedError()
        
